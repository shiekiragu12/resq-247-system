from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator

from .decorators import authorized_user, check_facility_and_attach_it_to_request
from .forms import CreateFacility
import openpyxl
from .models import *


# Create your views here.

@login_required(login_url='signin')
@authorized_user
@check_facility_and_attach_it_to_request
def dashboard(request, facility_id):
    context = {
        'patients_count': Patient.objects.filter(facility=request.facility).count(),
        'docs_count': Doctor.objects.filter(facilities__in=[request.facility]).count(),
        'appointments': Appointment.objects.filter(facility=request.facility).count(),
        'doctors': Doctor.objects.filter(facilities__in=[request.facility]).order_by('-id')[0:5],
        'patients': Patient.objects.filter(facility=request.facility).order_by('-id')[0:5],
    }

    return render(request, template_name='dashboard/pages/index.html', context=context)


def login(request, facility_id):
    return render(request, template_name='dashboard/pages/login.html', context={})


def register(request, facility_id):
    return render(request, template_name='dashboard/pages/register.html', context={})


def facilities(request):
    facilities_ = Facility.objects.all()
    paginator = Paginator(facilities_, 25)  # Show 25 contacts per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    nairobi = County.objects.get(id=1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "counties": County.objects.all()
    }
    return render(request, template_name="facilities/facilities.html", context=context)


def facilities_per_location(request, location_type, location_id, location_slug):

    if location_type == "county":
        county = get_object_or_404(County, id=location_id)
        facilities_ = Facility.objects.filter(county=county)
        paginator = Paginator(facilities_, 25)  # Show 25 contacts per page.

        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        context = {
            "page_obj": page_obj,
            "pages": paginator.page_range,
            "page_count": paginator.num_pages,
            "counties": County.objects.all(),
            "location_type": location_type,
            "county": county,
        }
        return render(request, template_name="facilities/facilities_county.html", context=context)
    else:
        constituency = get_object_or_404(Constituency, id=location_id)
        facilities_ = Facility.objects.filter(constituency=constituency)
        paginator = Paginator(facilities_, 25)  # Show 25 contacts per page.

        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        context = {
            "page_obj": page_obj,
            "pages": paginator.page_range,
            "page_count": paginator.num_pages,
            "counties": County.objects.all(),
            "location_type": location_type,
            "constituency": constituency,
        }
        return render(request, template_name="facilities/facilities_constituency.html", context=context)


def create_facility(request):
    if request.method == 'POST':
        request.POST._mutable = True
        request.POST.update({"owner": request.user.id})
        request.POST._mutable = False
        form = CreateFacility(request.POST, request.FILES)

        if form.is_valid():
            form.save()
            messages.success(request, 'Facility added successfully')
            return redirect('account_profile')
        else:
            for error in form.errors:
                messages.error(request, f"\n{error} {form.errors[error]}\n")
            return redirect('account_profile')


@login_required(login_url='signin')
@check_facility_and_attach_it_to_request
def facility(request, facility_id, facility_slug):
    return render(request, template_name='facilities/facility_home.html', context={})


@login_required(login_url='signin')
@check_facility_and_attach_it_to_request
def edit_facility(request, facility_id):
    if request.method == "POST":
        if request.POST.get('home_page_content', None) is not None:
            request.facility.home_page_content = request.POST.get('home_page_content')
            request.facility.save()
            messages.success(request, "Home page content updated successfully")
        if request.POST.get('about_page_content', None) is not None:
            request.facility.about_page_content = request.POST.get('about_page_content')
            request.facility.save()
            messages.success(request, "About page content updated successfully")
        if request.POST.get('online_page_content', None) is not None:
            request.facility.online_page_content = request.POST.get('online_page_content')
            request.facility.save()
            messages.success(request, "Online services page content updated successfully")

        return redirect('edit_facility', facility_id)
    return render(request, template_name='dashboard/pages/editfacility.html', context={})


@login_required(login_url='signin')
@check_facility_and_attach_it_to_request
def facility_about(request, facility_id, facility_slug):
    return render(request, template_name='facilities/facility_about.html', context={})


@login_required(login_url='signin')
@check_facility_and_attach_it_to_request
def facility_online_services(request, facility_id, facility_slug):
    return render(request, template_name='facilities/facility_online_services.html', context={})


@login_required(login_url='signin')
@check_facility_and_attach_it_to_request
def facility_services_treatments(request, facility_id, facility_slug):
    return render(request, template_name='facilities/facility_services_treatments.html', context={})


@login_required(login_url='signin')
@authorized_user
@check_facility_and_attach_it_to_request
def patients(request, facility_id):
    # messages.success(request, "almost done")
    context = {
        "patients": Patient.objects.filter(facility=request.facility)
    }
    return render(request, template_name='dashboard/pages/patient-list.html', context=context)


@login_required(login_url='signin')
@authorized_user
@check_facility_and_attach_it_to_request
def patients_create(request, facility_id):
    if request.method == "POST":
        existing_user = User.objects.filter(Q(username=get_value(request, 'email')) |
                                            Q(email=get_value(request, 'email'))).first()
        if existing_user is None:
            user = User.objects.create(username=get_value(request, 'email'),
                                       first_name=get_value(request, 'first_name'),
                                       last_name=get_value(request, 'last_name'),
                                       email=get_value(request, 'email'))
            user.save()
            user.set_password(get_value(request, 'username'))
            user.refresh_from_db()
            user.profile.profile_photo = request.FILES.get('profile_photo')
            user.profile.phone_number = get_value(request, 'phone_number')
            user.profile.gender = get_value(request, 'gender')
            user.profile.address = get_value(request, 'address')
            user.profile.city = get_value(request, 'city')
            user.profile.postal_code = get_value(request, 'postal_code')
            user.profile.save()

            user.refresh_from_db()

            patient = Patient.objects.create(facility=request.facility, user=user,
                                             dob=get_value(request, 'dob'),
                                             blood_group=get_value(request, 'blood_group'))
            patient.save()
            messages.success(request, 'Patient created successfully')

        else:
            messages.error(request, "Cannot register patient as user with similar records already exists.")

    return redirect('patients', facility_id)


def get_value(request, key):
    return request.POST.get(key)


@login_required(login_url='signin')
@authorized_user
@check_facility_and_attach_it_to_request
def doctors(request, facility_id):
    context = {
        "specialities": FacilitySpeciality.objects.all(),
        "doctors": Doctor.objects.filter(facilities__in=[request.facility])
    }
    return render(request, template_name='dashboard/pages/doctor-list.html', context=context)


@login_required(login_url='signin')
@authorized_user
@check_facility_and_attach_it_to_request
def doctors_create(request, facility_id):
    if request.method == "POST":

        existing_user = User.objects.filter(Q(username=get_value(request, 'email')) |
                                            Q(email=get_value(request, 'email'))).first()
        if existing_user is None:
            user = User.objects.create(username=get_value(request, 'email'),
                                       first_name=get_value(request, 'first_name'),
                                       last_name=get_value(request, 'last_name'),
                                       email=get_value(request, 'email'))
            user.save()
            user.set_password(get_value(request, 'username'))
            user.refresh_from_db()
            user.profile.profile_photo = request.FILES.get('profile_photo')
            user.profile.phone_number = get_value(request, 'phone_number')
            user.profile.gender = get_value(request, 'gender')
            user.profile.address = get_value(request, 'address')
            user.profile.city = get_value(request, 'city')
            user.profile.postal_code = get_value(request, 'postal_code')
            user.profile.save()

            user.refresh_from_db()

            speciality = FacilitySpeciality.objects.filter(id=get_value(request, 'speciality')).first()
            doctor = Doctor.objects.create(facility=request.facility, user=user,
                                           speciality=speciality,
                                           description=get_value(request, 'description'))
            doctor.save()
            messages.success(request, 'Doctor created successfully')

        else:
            messages.error(request, "Cannot register doctor as user with similar records already exists.")

    return redirect('doctors', facility_id)


@login_required(login_url='signin')
@authorized_user
@check_facility_and_attach_it_to_request
def staff(request, facility_id):
    context = {
        "staffs": Staff.objects.filter(facility=request.facility)
    }
    return render(request, template_name='dashboard/pages/staff.html', context=context)


@login_required(login_url='signin')
@authorized_user
@check_facility_and_attach_it_to_request
def staff_create(request, facility_id):
    if request.method == "POST":

        existing_user = User.objects.filter(Q(username=get_value(request, 'username')) |
                                            Q(email=get_value(request, 'email'))).first()
        if existing_user is None:
            user = User.objects.create(username=get_value(request, 'email'),
                                       first_name=get_value(request, 'first_name'),
                                       last_name=get_value(request, 'last_name'),
                                       email=get_value(request, 'email'))
            user.save()
            user.set_password(get_value(request, 'username'))
            user.refresh_from_db()
            user.profile.profile_photo = request.FILES.get('profile_photo')
            user.profile.phone_number = get_value(request, 'phone_number')
            user.profile.gender = get_value(request, 'gender')
            user.profile.dob = get_value(request, 'dob')
            user.profile.address = get_value(request, 'address')
            user.profile.city = get_value(request, 'city')
            user.profile.postal_code = get_value(request, 'postal_code')
            user.profile.save()

            user.refresh_from_db()

            staffMember = Staff.objects.create(facility=request.facility, user=user,
                                               designation=get_value(request, 'designation'),
                                               education=get_value(request, 'education'))
            staffMember.save()
            messages.success(request, 'Staff created successfully')

        else:
            messages.error(request, "Cannot register staff as user with similar records already exists.")

    return redirect('staff', facility_id)


@login_required(login_url='signin')
@authorized_user
@check_facility_and_attach_it_to_request
def appointments(request, facility_id):
    context = {
        'doctors': Doctor.objects.filter(facility=request.facility),
        'patients': Patient.objects.filter(facility=request.facility),
        'conditions': Condition.objects.all(),
        'appointments': Appointment.objects.filter(facility=request.facility)
    }
    return render(request, template_name='dashboard/pages/appointment.html', context=context)


@login_required(login_url='signin')
@authorized_user
@check_facility_and_attach_it_to_request
def appointments_create(request, facility_id):
    if request.method == "POST":
        patient = Patient.objects.filter(id=get_value(request, 'patient')).first()
        doctor = Doctor.objects.filter(id=get_value(request, 'doctor')).first()
        condition_ = Condition.objects.filter(id=get_value(request, 'condition')).first()

        appointment = Appointment.objects.create(facility=request.facility, doctor=doctor,
                                                 patient=patient, note=get_value(request, 'note'),
                                                 start_time=get_value(request, 'start_time'),
                                                 end_time=get_value(request, 'end_time'),
                                                 condition=condition_, date=get_value(request, 'date'))
        appointment.save()
        messages.success(request, "Appointment created successfully")
    return redirect('appointments', facility_id)


@login_required(login_url='signin')
@authorized_user
@check_facility_and_attach_it_to_request
def services(request, facility_id):
    return render(request, template_name='dashboard/pages/services.html', context={})


@login_required(login_url='signin')
@authorized_user
@check_facility_and_attach_it_to_request
def encounters(request, facility_id):
    return render(request, template_name='dashboard/pages/encounters.html', context={})


def upload_conditions(request):
    if request.method == 'POST':
        excel_file = request.FILES.get("file")
        wb = openpyxl.load_workbook(excel_file)
        excel_data = list()

        for letter in ["A", "B"]:
            worksheet = wb[letter]
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)

        # print(excel_data)

        conditions = []
        for condition_data in excel_data:
            available = Condition.objects.filter(name=condition_data[0]).first()
            if available is None:
                cond = Condition(name=condition_data[0],
                                 factsheet=condition_data[1],
                                 pathogen=condition_data[2],
                                 clinical_features=condition_data[3],
                                 transmission=condition_data[4],
                                 diagnosis=condition_data[5],
                                 treatment=condition_data[6],
                                 prevention=condition_data[7])
                conditions.append(cond)
        Condition.objects.bulk_create(conditions)

    return render(request, template_name='health/uploadconditions.html', context={})


def upload_medical_facilities(request):
    if request.method == 'POST':
        excel_file = request.FILES.get("file")
        wb = openpyxl.load_workbook(excel_file)
        print(wb)
        excel_data = list()

        for county in ["A", "B", "C", "D", "F", "G", "H"]:
            worksheet = wb[county]
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)

        print(excel_data)

        facilities = []
        for facility_data in excel_data:
            available = Facility.objects.filter(name=facility_data[0]).first()
            county_ = County.objects.filter(name__iexact=facility_data[3]).first()
            county = None
            if county_ is None:
                county = County.objects.create(name=facility_data[3])
            consti_ = Constituency.objects.filter(name__iexact=facility_data[4]).first()
            consti = None
            if consti_ is None:
                consti = Constituency.objects.create(name=facility_data[4])

            f_type_ = FacilityType.objects.filter(name__iexact=facility_data[1]).first()
            f_type = None
            if f_type_ is None:
                f_type = FacilityType.objects.create(name=facility_data[1])

            if available is None:
                facility_ = Facility(owner_name=facility_data[2],
                                     facility_type=f_type if f_type_ is None else f_type_,
                                     name=facility_data[0],
                                     county=county if county_ is None else county_,
                                     constituency=consti if consti_ is None else consti_,
                                     latitude=facility_data[5],
                                     longitude=facility_data[6],
                                     email=facility_data[7],
                                     contact_no=facility_data[8])
                facilities.append(facility_)
        Facility.objects.bulk_create(facilities)

    return render(request, template_name='health/uploadfacilities.html', context={})
